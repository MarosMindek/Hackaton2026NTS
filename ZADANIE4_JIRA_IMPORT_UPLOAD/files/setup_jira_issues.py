#!/usr/bin/env python3
"""
setup_jira_issues.py
Creates Jira issues from Excel timesheet data so timesheet_tool.py can log time against them.

Usage:
  python setup_jira_issues.py --project MYPROJECT
  python setup_jira_issues.py --project MYPROJECT --dry-run
  python setup_jira_issues.py --list-projects
"""

import sys
import json
import os
import argparse
from pathlib import Path

import requests
import urllib3
import openpyxl

BASE_DIR = Path(__file__).parent

if sys.platform == "win32":
    os.system("")
_USE_COLOR = sys.stdout.isatty()


def _c(code: str) -> str:
    return code if _USE_COLOR else ""


C_OK   = "\033[32m"
C_WARN = "\033[33m"
C_ERR  = "\033[31m"
C_RST  = "\033[0m"


# ── Config ─────────────────────────────────────────────────────────────────────
def load_config() -> dict:
    p = BASE_DIR / "config.json"
    if not p.exists():
        print("[!] config.json not found. Run timesheet_tool.py first to generate it.")
        sys.exit(1)
    cfg = json.loads(p.read_text(encoding="utf-8"))
    if cfg.get("jira_api_token", "").startswith("YOUR_"):
        print("[!] config.json still has placeholder values.")
        print("    Fill in jira_url, jira_email, jira_api_token before running this script.")
        sys.exit(1)
    return cfg


# ── Jira API helpers ───────────────────────────────────────────────────────────
class JiraClient:
    def __init__(self, config: dict):
        self.base  = config["jira_url"].rstrip("/")
        self.auth  = (config["jira_email"], config["jira_api_token"])
        self.hdrs  = {"Accept": "application/json", "Content-Type": "application/json"}
        self.session = requests.Session()
        self.session.verify = config.get("ssl_verify", True)
        if not self.session.verify:
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    def get(self, path: str, params: dict = None) -> requests.Response:
        return self.session.get(f"{self.base}{path}", auth=self.auth,
                                headers=self.hdrs, params=params, timeout=15)

    def post(self, path: str, body: dict) -> requests.Response:
        return self.session.post(f"{self.base}{path}", auth=self.auth,
                                 headers=self.hdrs, json=body, timeout=15)

    def myself(self) -> dict:
        r = self.get("/rest/api/3/myself")
        r.raise_for_status()
        return r.json()

    def get_project(self, key: str) -> dict | None:
        r = self.get(f"/rest/api/3/project/{key}")
        if r.status_code == 404:
            return None
        r.raise_for_status()
        return r.json()

    def list_projects(self) -> list[dict]:
        r = self.get("/rest/api/3/project/search", params={"maxResults": 50})
        r.raise_for_status()
        return r.json().get("values", [])

    def get_issue_types(self, project_key: str) -> list[dict]:
        r = self.get(f"/rest/api/3/project/{project_key}")
        r.raise_for_status()
        return r.json().get("issueTypes", [])

    def create_issue(self, project_key: str, summary: str,
                     issue_type: str, account_id: str) -> dict:
        body = {
            "fields": {
                "project":   {"key": project_key},
                "summary":   summary,
                "issuetype": {"name": issue_type},
                "assignee":  {"accountId": account_id},
            }
        }
        r = self.post("/rest/api/3/issue", body)
        if r.status_code == 400:
            # Retry without assignee (some projects disallow it)
            del body["fields"]["assignee"]
            r = self.post("/rest/api/3/issue", body)
        r.raise_for_status()
        return r.json()


# ── Excel extraction ───────────────────────────────────────────────────────────
def extract_jira_issues(path: Path) -> dict[str, dict]:
    """Returns {old_key: {summary, issue_type}} for unique issues."""
    wb = openpyxl.load_workbook(path)
    ws = wb["Worklogs"]
    headers = [str(c.value or "").strip() for c in ws[1]]
    hi = {h: i for i, h in enumerate(headers)}

    issues: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key     = str(row[hi.get("Issue Key", -1)] or "").strip()
        summary = str(row[hi.get("Issue summary", -1)] or "").strip()
        itype   = str(row[hi.get("Issue Type", -1)] or "Task").strip() or "Task"
        if key and key not in issues:
            issues[key] = {"summary": summary, "issue_type": itype}
    return issues


def extract_rwe_worklists(path: Path) -> dict[str, dict]:
    """Returns {worklist_name: {wbs, summary}} for unique RWE worklists."""
    wb = openpyxl.load_workbook(path)
    ws = wb["SAPUI5 Export"]
    headers = [str(c.value or "").strip() for c in ws[1]]
    hi = {h: i for i, h in enumerate(headers)}

    worklists: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        worklist = str(row[hi.get("Worklist Name", -1)] or "").strip()
        wbs      = str(row[hi.get("WBS Element", -1)] or "").strip()
        if worklist and worklist not in worklists:
            worklists[worklist] = {
                "wbs": wbs,
                "summary": f"RWE: {worklist}",
                "issue_type": "Task",
            }
    return worklists


# ── Pick best available issue type ────────────────────────────────────────────
def pick_issue_type(wanted: str, available: list[dict]) -> str:
    names = [t["name"] for t in available]
    if wanted in names:
        return wanted
    for fallback in ("Task", "Story", "Bug", "Sub-task"):
        if fallback in names:
            return fallback
    return names[0] if names else "Task"


# ── Commands ───────────────────────────────────────────────────────────────────
def cmd_list_projects(jira: JiraClient) -> None:
    print("\nAvailable Jira projects:\n")
    projects = jira.list_projects()
    if not projects:
        print("  (none found — check your permissions)")
        return
    for p in projects:
        print(f"  {p['key']:<15} {p['name']}")
    print(f"\n  Total: {len(projects)}")
    print("\nUse one of these keys with:  python setup_jira_issues.py --project <KEY>")


def cmd_create_issues(jira: JiraClient, config: dict, project_key: str,
                      dry_run: bool) -> None:
    # Verify project exists
    project = jira.get_project(project_key)
    if not project:
        print(f"[!] Project '{project_key}' not found in your Jira.")
        print(f"    Run --list-projects to see available projects.")
        sys.exit(1)

    print(f"\n  Project : {project['name']} ({project_key})")

    # Get available issue types for this project
    issue_types = jira.get_issue_types(project_key)
    type_names  = [t["name"] for t in issue_types]
    print(f"  Issue types available: {', '.join(type_names)}")

    # Get current user accountId
    me = jira.myself()
    account_id = me["accountId"]
    print(f"  Creating as: {me.get('displayName')} ({account_id})\n")

    # Collect issues to create
    jira_issues: dict[str, dict] = {}
    rwe_worklists: dict[str, dict] = {}

    jira_xlsx = BASE_DIR / "JIRA EXPORT.xlsx"
    rwe_xlsx  = BASE_DIR / "RWE export.xlsx"

    if jira_xlsx.exists():
        jira_issues = extract_jira_issues(jira_xlsx)
        print(f"  JIRA EXPORT.xlsx — {len(jira_issues)} unique issues")
    if rwe_xlsx.exists():
        rwe_worklists = extract_rwe_worklists(rwe_xlsx)
        print(f"  RWE export.xlsx  — {len(rwe_worklists)} unique worklists\n")

    if not jira_issues and not rwe_worklists:
        print("[!] No issues found in Excel files.")
        sys.exit(0)

    label = "[DRY RUN] " if dry_run else ""
    print(f"  {label}Will create {len(jira_issues) + len(rwe_worklists)} issue(s) in {project_key}:\n")

    # Show preview table
    all_items = (
        [(k, v["summary"], v["issue_type"], "JIRA_EXPORT") for k, v in jira_issues.items()] +
        [(wl, v["summary"], v["issue_type"], "RWE_EXPORT")  for wl, v in rwe_worklists.items()]
    )
    for old_key, summary, itype, src in all_items:
        print(f"  {src:<12}  {old_key:<22}  [{itype}]  {summary[:55]}")

    print()
    if not dry_run:
        confirm = input(f"  Create {len(all_items)} issue(s) in project {project_key}? [y/N]: ").strip().lower()
        if confirm != "y":
            print("  Cancelled.")
            sys.exit(0)

    # Create issues and build mapping
    key_mapping: dict[str, str]      = {}  # old_jira_key  -> new_jira_key
    wbs_mapping: dict[str, str]      = {}  # wbs_element   -> new_jira_key
    worklist_mapping: dict[str, str] = {}  # worklist_name -> new_jira_key

    ok = 0
    failed = 0

    print()
    for old_key, meta in jira_issues.items():
        itype   = pick_issue_type(meta["issue_type"], issue_types)
        # Embed original key in summary so it's identifiable in Jira
        summary = f"[{old_key}] {meta['summary']}"
        if dry_run:
            print(f"  {_c(C_OK)}[DRY]{_c(C_RST)} {old_key} → would create '{summary[:60]}' as {itype}")
            key_mapping[old_key] = f"{project_key}-???"
            continue
        try:
            result  = jira.create_issue(project_key, summary, itype, account_id)
            new_key = result["key"]
            key_mapping[old_key] = new_key
            print(f"  {_c(C_OK)}[OK]{_c(C_RST)}  {old_key} → {new_key}  ({summary[:55]})")
            ok += 1
        except Exception as exc:
            print(f"  {_c(C_ERR)}[FAIL]{_c(C_RST)} {old_key} → {exc}")
            failed += 1

    for worklist, meta in rwe_worklists.items():
        itype   = pick_issue_type(meta["issue_type"], issue_types)
        summary = f"[RWE] {meta['summary']}"
        if dry_run:
            print(f"  {_c(C_OK)}[DRY]{_c(C_RST)} RWE/{worklist[:35]} → would create '{summary[:50]}' as {itype}")
            worklist_mapping[worklist] = f"{project_key}-???"
            wbs_mapping[meta["wbs"]]   = f"{project_key}-???"
            continue
        try:
            result  = jira.create_issue(project_key, summary, itype, account_id)
            new_key = result["key"]
            worklist_mapping[worklist] = new_key
            if meta["wbs"]:
                wbs_mapping[meta["wbs"]] = new_key
            print(f"  {_c(C_OK)}[OK]{_c(C_RST)}  RWE/{worklist[:35]} → {new_key}  ({summary[:45]})")
            ok += 1
        except Exception as exc:
            print(f"  {_c(C_ERR)}[FAIL]{_c(C_RST)} RWE/{worklist[:35]} → {exc}")
            failed += 1

    # ── Report ──────────────────────────────────────────────────────────────────
    print(f"\n  {'=' * 50}")
    if dry_run:
        print(f"  DRY RUN complete — no issues were created.")
        print(f"  Remove --dry-run to actually create them.")
        return

    print(f"  Created: {ok}   Failed: {failed}")

    if ok == 0:
        print("\n  [!] Nothing was created. Check project permissions.")
        return

    # ── Save full key map ────────────────────────────────────────────────────────
    map_path = BASE_DIR / "key_mapping.json"
    full_map = {
        "_comment": "Maps old Excel issue keys / WBS / Worklist to new Jira keys",
        "key_mapping":      key_mapping,
        "wbs_mapping":      wbs_mapping,
        "worklist_mapping": worklist_mapping,
    }
    map_path.write_text(json.dumps(full_map, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\n  Key mapping saved to: {map_path}")

    # ── Patch mapping.json used by timesheet_tool.py ─────────────────────────────
    mapping_json = BASE_DIR / "mapping.json"
    if mapping_json.exists():
        mapping = json.loads(mapping_json.read_text(encoding="utf-8"))
    else:
        mapping = {"wbs_to_issue": {}, "worklist_to_issue": {}}

    mapping["wbs_to_issue"].update(wbs_mapping)
    mapping["worklist_to_issue"].update(worklist_mapping)
    mapping_json.write_text(json.dumps(mapping, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  mapping.json updated with RWE → Jira mappings")

    # ── Patch config.json author_account_id ─────────────────────────────────────
    cfg_path = BASE_DIR / "config.json"
    cfg      = json.loads(cfg_path.read_text(encoding="utf-8"))
    if cfg.get("author_account_id", "").startswith("YOUR_"):
        cfg["author_account_id"] = account_id
        cfg_path.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"  config.json: author_account_id set to {account_id}")

    # ── Offer to launch timesheet_tool immediately ───────────────────────────────
    print(f"\n  {'=' * 50}")
    print(f"  All done! Issues created and key_mapping.json saved.")
    print(f"\n  Run timesheet tool now? It will use the new Jira keys automatically.")
    print(f"  Command: python -X utf8 timesheet_tool.py --remap key_mapping.json --dry-run")
    choice = input("\n  Launch timesheet_tool.py --dry-run now? [y/N]: ").strip().lower()
    if choice == "y":
        import subprocess
        subprocess.run(
            [sys.executable, "-X", "utf8", str(BASE_DIR / "timesheet_tool.py"),
             "--remap", str(map_path), "--dry-run"],
            cwd=str(BASE_DIR),
        )


# ── Main ────────────────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Create Jira issues from Excel timesheet data.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python setup_jira_issues.py --list-projects
  python setup_jira_issues.py --project MYPROJ --dry-run
  python setup_jira_issues.py --project MYPROJ
"""
    )
    parser.add_argument("--project",       help="Jira project key to create issues in (e.g. MYPROJ)")
    parser.add_argument("--list-projects", action="store_true", help="List all accessible Jira projects and exit")
    parser.add_argument("--dry-run",       action="store_true", help="Show what would be created without actually creating")
    parser.add_argument("--config",        default="config.json", help="Path to config.json")
    args = parser.parse_args()

    config = load_config()
    jira   = JiraClient(config)

    # Verify connectivity
    try:
        me = jira.myself()
        print(f"[*] Connected to Jira as: {me.get('displayName')} ({me.get('emailAddress')})")
    except Exception as exc:
        print(f"[!] Cannot connect to Jira: {exc}")
        print(f"    Check jira_url, jira_email, jira_api_token in config.json")
        sys.exit(1)

    if args.list_projects:
        cmd_list_projects(jira)
        return

    if not args.project:
        print("[!] Specify a project with --project KEY, or use --list-projects to see options.")
        sys.exit(1)

    cmd_create_issues(jira, config, args.project.upper(), dry_run=args.dry_run)


if __name__ == "__main__":
    main()
