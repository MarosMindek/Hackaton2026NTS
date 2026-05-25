#!/usr/bin/env python3
"""
Timesheet Converter: Excel -> XML -> Tempo Timesheets
Supports: JIRA EXPORT.xlsx, RWE export.xlsx
"""

import sys
import json
import csv
import re
import os
import argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Optional

import requests
import urllib3
import openpyxl
from lxml import etree

# Module-level session — configured once in main() after config is loaded
_session = requests.Session()

# ── Constants ──────────────────────────────────────────────────────────────────
MIN_HOURS = 0.25        # 15 minutes
MAX_HOURS_PER_DAY = 24.0
SECONDS_PER_HOUR = 3600
BASE_DIR = Path(__file__).parent

# Enable ANSI color codes on Windows terminal
if sys.platform == "win32":
    os.system("")  # activates VT100 in Windows Console

# Disable colors if stdout is not a real terminal (piped/redirected)
_USE_COLOR = sys.stdout.isatty()

# ── Config ─────────────────────────────────────────────────────────────────────
DEFAULT_CONFIG = {
    "jira_url": "https://your-company.atlassian.net",
    "jira_email": "your.email@company.com",
    "jira_api_token": "YOUR_JIRA_API_TOKEN",
    "tempo_token": "YOUR_TEMPO_TOKEN",
    "author_account_id": "YOUR_JIRA_ACCOUNT_ID",
    "default_start_time": "09:00:00"
}

DEFAULT_MAPPING = {
    "_comment": "Map WBS Element or Worklist Name to a Jira Issue Key",
    "wbs_to_issue": {
        "1300-00529GAA006272914": "",
        "1300-00168GAA005082506": "",
        "1300-00168GAA005082507": "",
        "1300-00529GAA006272907": ""
    },
    "worklist_to_issue": {
        "CITS/TS4 POWER MSF  RE-FX": "",
        "CITS/Rx4 add.Serv. RWEST": "",
        "CITS/Rx4 add.Serv. RWETI": "",
        "TS4_POWER_MSF_Controlling (Hypercare)": ""
    }
}


def load_config(path: Path) -> dict:
    if not path.exists():
        path.write_text(json.dumps(DEFAULT_CONFIG, indent=2), encoding="utf-8")
        print(f"[INFO] Created default config: {path}")
        print("[!]  Fill in your credentials in config.json and re-run.")
        sys.exit(0)
    return json.loads(path.read_text(encoding="utf-8"))


def load_mapping(path: Path) -> dict:
    if not path.exists():
        path.write_text(json.dumps(DEFAULT_MAPPING, indent=2), encoding="utf-8")
        print(f"[INFO] Created mapping template: {path}")
        print("[!]  Edit mapping.json — add Jira Issue Keys for each WBS/Worklist entry.")
    return json.loads(path.read_text(encoding="utf-8"))


def config_is_placeholder(config: dict) -> bool:
    return (
        config.get("jira_api_token", "").startswith("YOUR_")
        or config.get("tempo_token", "").startswith("YOUR_")
    )


# ── Date / Time helpers ────────────────────────────────────────────────────────
def parse_date(val) -> Optional[str]:
    """Return ISO date string YYYY-MM-DD, or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def parse_hours(val) -> Optional[float]:
    """Return hours as float, or None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    m = re.match(r"^(\d+):(\d{2})$", s)
    if m:
        return int(m.group(1)) + int(m.group(2)) / 60.0
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return None


# ── Excel parsers ──────────────────────────────────────────────────────────────
def parse_jira_export(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb["Worklogs"]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        r = dict(zip(headers, row))
        entries.append({
            "source": "JIRA_EXPORT",
            "issue_key": str(r.get("Issue Key") or "").strip(),
            "description": str(r.get("Work Description") or "").strip(),
            "date": parse_date(r.get("Work date")),
            "hours": parse_hours(r.get("Hours")),
            "author": str(r.get("Username") or "").strip(),
            "full_name": str(r.get("Full name") or "").strip(),
            "project_key": str(r.get("Project Key") or "").strip(),
            "account_key": str(r.get("Account Key") or "").strip(),
            "wbs_element": "",
            "worklist_name": "",
        })
    return entries


def parse_rwe_export(path: Path, mapping: dict) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb["SAPUI5 Export"]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

    wbs_map = mapping.get("wbs_to_issue", {})
    worklist_map = mapping.get("worklist_to_issue", {})

    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        r = dict(zip(headers, row))

        wbs = str(r.get("WBS Element") or "").strip()
        worklist = str(r.get("Worklist Name") or "").strip()
        desc = str(r.get("Longtext Detail") or "").strip()

        issue_key = wbs_map.get(wbs) or worklist_map.get(worklist) or ""

        entries.append({
            "source": "RWE_EXPORT",
            "issue_key": issue_key,
            "description": desc,
            "date": parse_date(r.get("Date")),
            "hours": parse_hours(r.get("Hours")),
            "author": str(r.get("Employee/app.name") or "").strip(),
            "full_name": str(r.get("Employee/app.name") or "").strip(),
            "project_key": "",
            "account_key": "",
            "wbs_element": wbs,
            "worklist_name": worklist,
            "status_text": str(r.get("Status Text") or "").strip(),
        })
    return entries


# ── Jira API ───────────────────────────────────────────────────────────────────
def validate_issues_via_jira(issue_keys: set, config: dict) -> dict[str, bool]:
    """Returns {issue_key: exists_in_jira}."""
    results: dict[str, bool] = {}
    base = config["jira_url"].rstrip("/")
    auth = (config["jira_email"], config["jira_api_token"])
    headers = {"Accept": "application/json"}

    for key in sorted(issue_keys):
        if not key:
            continue
        try:
            resp = _session.get(
                f"{base}/rest/api/3/issue/{key}",
                auth=auth, headers=headers, timeout=10
            )
            results[key] = resp.status_code == 200
            status = "OK" if results[key] else f"NOT FOUND (HTTP {resp.status_code})"
            print(f"  {key}: {status}")
        except Exception as exc:
            results[key] = False
            print(f"  {key}: ERROR — {exc}")

    return results


def resolve_issue_ids(issue_keys: set, config: dict) -> dict[str, int]:
    """Returns {issue_key: numeric_jira_id} needed by Tempo API."""
    result: dict[str, int] = {}
    base = config["jira_url"].rstrip("/")
    auth = (config["jira_email"], config["jira_api_token"])
    headers = {"Accept": "application/json"}

    for key in sorted(issue_keys):
        if not key:
            continue
        try:
            resp = _session.get(
                f"{base}/rest/api/3/issue/{key}",
                auth=auth, headers=headers, timeout=10,
            )
            if resp.status_code == 200:
                numeric_id = int(resp.json()["id"])
                result[key] = numeric_id
            else:
                print(f"  [!] Could not resolve ID for {key}: HTTP {resp.status_code}")
        except Exception as exc:
            print(f"  [!] Could not resolve ID for {key}: {exc}")

    return result


def get_account_id_from_jira(config: dict, email: str) -> Optional[str]:
    """Resolve a Jira accountId from email address."""
    try:
        base = config["jira_url"].rstrip("/")
        auth = (config["jira_email"], config["jira_api_token"])
        resp = _session.get(
            f"{base}/rest/api/3/user/search",
            params={"query": email},
            auth=auth,
            headers={"Accept": "application/json"},
            timeout=10
        )
        if resp.status_code == 200:
            users = resp.json()
            if users:
                return users[0].get("accountId")
    except Exception:
        pass
    return None


# ── Validation ─────────────────────────────────────────────────────────────────
def validate_entry(entry: dict, issue_validity: dict) -> tuple[str, list[str]]:
    """Return ('OK'|'Warning'|'Error', [messages])."""
    errors: list[str] = []
    warnings: list[str] = []

    ik = entry.get("issue_key", "")
    if not ik:
        if entry["source"] == "RWE_EXPORT":
            warnings.append(
                f"No Issue Key mapped — WBS={entry.get('wbs_element') or 'n/a'} "
                f"/ Worklist={entry.get('worklist_name') or 'n/a'}"
            )
        else:
            errors.append("Missing Issue Key")
    elif ik in issue_validity and not issue_validity[ik]:
        errors.append(f"Issue '{ik}' not found or inaccessible in Jira")

    if not entry.get("date"):
        errors.append("Missing or unparseable date")

    hours = entry.get("hours")
    if hours is None:
        errors.append("Missing hours value")
    elif hours <= 0:
        errors.append(f"Hours must be > 0 (got {hours})")
    elif hours > MAX_HOURS_PER_DAY:
        errors.append(f"Hours {hours} exceeds 24h/day limit")
    elif hours < MIN_HOURS:
        errors.append(f"Hours {hours} is below 15-minute minimum (0.25h)")
    elif round(hours % 0.25, 4) != 0:
        warnings.append(f"Hours {hours} is not aligned to 15-minute granularity — will be rounded")

    if errors:
        return "Error", errors + warnings
    if warnings:
        return "Warning", warnings
    return "OK", []


# ── XML builder ────────────────────────────────────────────────────────────────
def build_xml(entries: list[dict], issue_validity: dict, config: dict,
              issue_id_map: dict[str, int] | None = None) -> etree._Element:
    root = etree.Element("worklogs", generated=datetime.now().isoformat())

    # Detect duplicate (issueKey, date) pairs
    key_counts: dict = defaultdict(list)
    for i, e in enumerate(entries):
        key_counts[(e.get("issue_key"), e.get("date"))].append(i)
    duplicate_indices = {i for indices in key_counts.values() if len(indices) > 1 for i in indices}

    for i, entry in enumerate(entries):
        status, messages = validate_entry(entry, issue_validity)

        if i in duplicate_indices and entry.get("issue_key"):
            if status == "OK":
                status = "Warning"
            if not any("uplicate" in m for m in messages):
                messages.append("Duplicate: same Issue Key + date detected")

        hours = entry.get("hours") or 0.0
        # Snap to 15-min boundary
        snapped = round(hours * 4) / 4
        seconds = int(snapped * SECONDS_PER_HOUR)

        wl = etree.SubElement(root, "worklog",
            id=str(i + 1),
            source=entry.get("source", ""),
            status=status,
        )

        def add(tag: str, val) -> None:
            el = etree.SubElement(wl, tag)
            el.text = str(val) if val is not None else ""

        ik = entry.get("issue_key", "")
        add("issueKey",          ik)
        add("issueId",           str((issue_id_map or {}).get(ik, "")))
        add("date",              entry.get("date", ""))
        add("hours",             f"{snapped:.2f}")
        add("timeSpentSeconds",  str(seconds))
        add("startTime",         config.get("default_start_time", "09:00:00"))
        add("description",       entry.get("description", ""))
        add("author",            entry.get("author", ""))
        add("fullName",          entry.get("full_name", ""))
        add("authorAccountId",   config.get("author_account_id", ""))
        add("projectKey",        entry.get("project_key", ""))
        add("accountKey",        entry.get("account_key", ""))
        add("wbsElement",        entry.get("wbs_element", ""))
        add("worklistName",      entry.get("worklist_name", ""))

        val_el = etree.SubElement(wl, "validation")
        for msg in (messages or ["OK"]):
            etree.SubElement(val_el, "message").text = msg

    return root


# ── CLI review ─────────────────────────────────────────────────────────────────
def _c(code: str) -> str:
    return code if _USE_COLOR else ""

COLOR = {
    "OK":      "\033[32m",
    "Warning": "\033[33m",
    "Error":   "\033[31m",
    "reset":   "\033[0m",
}

STATUS_PREFIX = {
    "OK":      "   OK",
    "Warning": " WARN",
    "Error":   "  ERR",
}


def get_validation_messages(wl: etree._Element) -> list[str]:
    val = wl.find("validation")
    if val is None:
        return []
    return [m.text for m in val if m.text and m.text != "OK"]


def print_worklogs(worklogs: list[etree._Element]) -> None:
    header = f"{'ID':>4}  {'St.':<5}  {'Source':<12}  {'Issue Key':<22}  {'Date':<12}  {'Hrs':>5}  {'Author':<22}  Description"
    print("\n" + header)
    print("-" * 112)

    for wl in worklogs:
        sid   = wl.get("id", "")
        stat  = wl.get("status", "")
        src   = (wl.get("source") or "")[:12]
        ik    = (wl.findtext("issueKey") or "")[:22]
        dt    = (wl.findtext("date") or "")[:12]
        hrs   = wl.findtext("hours") or ""
        auth  = (wl.findtext("author") or wl.findtext("fullName") or "")[:22]
        desc  = (wl.findtext("description") or "")[:40]
        c     = _c(COLOR.get(stat, ""))
        r     = _c(COLOR["reset"])
        prefix = STATUS_PREFIX.get(stat, stat[:5])
        print(f"{sid:>4}  {c}{prefix}{r}  {src:<12}  {ik:<22}  {dt:<12}  {hrs:>5}  {auth:<22}  {desc}")

        if stat in ("Warning", "Error"):
            for msg in get_validation_messages(wl):
                print(f"             {c}[!] {msg}{r}")

    stats = [wl.get("status") for wl in worklogs]
    n_ok   = stats.count("OK")
    n_warn = stats.count("Warning")
    n_err  = stats.count("Error")
    print(f"\n  Total: {len(worklogs)}  |  OK: {n_ok}  |  WARN: {n_warn}  |  ERR: {n_err}")
    print(f"\n  Legend:")
    print(f"    {_c(COLOR['OK'])}   OK{_c(COLOR['reset'])}  — valid, will be submitted")
    print(f"    {_c(COLOR['Warning'])} WARN{_c(COLOR['reset'])}  — will be submitted, but has issues (see [!] lines) — verify before sending")
    print(f"    {_c(COLOR['Error'])}  ERR{_c(COLOR['reset'])}  — BLOCKED, will NOT be submitted — fix with e<id>")


def show_detail(wl: etree._Element) -> None:
    print(f"\n-- Worklog #{wl.get('id')} ---------------------------------------------------")
    for child in wl:
        if child.tag == "validation":
            for msg in child:
                print(f"  [validation] {msg.text}")
        else:
            print(f"  {child.tag}: {child.text or ''}")


def edit_worklog(wl: etree._Element) -> None:
    editable = {
        "1": "issueKey",
        "2": "date",
        "3": "hours",
        "4": "description",
        "5": "startTime",
        "6": "authorAccountId",
    }
    print(f"\n-- Edit Worklog #{wl.get('id')} --")
    for k, tag in editable.items():
        el = wl.find(tag)
        cur = el.text if el is not None else ""
        print(f"  {k}. {tag}: {cur}")

    choice = input("Field number to edit (or Enter to cancel): ").strip()
    tag = editable.get(choice)
    if not tag:
        print("  Cancelled.")
        return

    el = wl.find(tag)
    cur = el.text if el is not None else ""
    new_val = input(f"  New value [{cur}]: ").strip()
    if not new_val:
        print("  No change.")
        return

    if el is None:
        el = etree.SubElement(wl, tag)
    el.text = new_val

    # Keep timeSpentSeconds in sync
    if tag == "hours":
        try:
            h = float(new_val)
            snapped = round(h * 4) / 4
            el.text = f"{snapped:.2f}"
            ts = wl.find("timeSpentSeconds")
            if ts is not None:
                ts.text = str(int(snapped * SECONDS_PER_HOUR))
        except ValueError:
            pass

    # Update status after edit
    wl.set("status", "Edited — re-run validation")
    print(f"  [OK] Updated {tag}.")


def merge_duplicates(root: etree._Element) -> int:
    """Merge entries sharing (issueKey, date): sum hours, join descriptions."""
    worklogs = root.findall("worklog")
    groups: dict = defaultdict(list)
    for wl in worklogs:
        k = (wl.findtext("issueKey"), wl.findtext("date"))
        groups[k].append(wl)

    merged = 0
    for (ik, dt), group in groups.items():
        if len(group) < 2 or not ik:
            continue
        total_h = sum(float(wl.findtext("hours") or 0) for wl in group)
        descs = list(dict.fromkeys(
            wl.findtext("description") or "" for wl in group
        ))
        merged_desc = " | ".join(d for d in descs if d)

        keeper = group[0]
        keeper.find("hours").text = f"{total_h:.2f}"
        keeper.find("timeSpentSeconds").text = str(int(total_h * SECONDS_PER_HOUR))
        keeper.find("description").text = merged_desc
        keeper.set("status", "OK")
        # Clear duplicate warning
        val = keeper.find("validation")
        if val is not None:
            for m in list(val):
                val.remove(m)
            etree.SubElement(val, "message").text = f"Merged {len(group)} entries"

        for wl in group[1:]:
            root.remove(wl)
        merged += len(group) - 1

    return merged


def print_issues_summary(worklogs: list[etree._Element]) -> None:
    """Print a grouped summary of all warnings and errors before submit."""
    c_warn = _c(COLOR["Warning"])
    c_err  = _c(COLOR["Error"])
    r      = _c(COLOR["reset"])

    warn_entries = [(wl, get_validation_messages(wl)) for wl in worklogs
                    if wl.get("status") == "Warning" and get_validation_messages(wl)]
    err_entries  = [(wl, get_validation_messages(wl)) for wl in worklogs
                    if wl.get("status") == "Error"]

    if not warn_entries and not err_entries:
        print(f"\n  {_c(COLOR['OK'])}All entries are valid.{r}")
        return

    print(f"\n{'=' * 60}")
    print(" ISSUES FOUND — read before submitting")
    print(f"{'=' * 60}")

    # Group warnings by message text for compact view
    warn_groups: dict[str, list[str]] = defaultdict(list)
    for wl, msgs in warn_entries:
        for msg in msgs:
            warn_groups[msg].append(f"#{wl.get('id')} {wl.findtext('issueKey') or '(no key)'} {wl.findtext('date') or ''}")

    if warn_groups:
        print(f"\n  {c_warn}WARNINGS (will be submitted, but check these):{r}")
        for reason, ids in warn_groups.items():
            print(f"\n    {c_warn}[!] {reason}{r}")
            # Show at most 8 entries per group, then summarize
            for entry in ids[:8]:
                print(f"        {entry}")
            if len(ids) > 8:
                print(f"        ... and {len(ids) - 8} more")

    if err_entries:
        print(f"\n  {c_err}ERRORS (will NOT be submitted — fix with e<id>):{r}")
        for wl, msgs in err_entries:
            ik = wl.findtext("issueKey") or "(no key)"
            dt = wl.findtext("date") or ""
            print(f"\n    {c_err}#{wl.get('id')} {ik} {dt}{r}")
            for msg in msgs:
                print(f"        {c_err}[!] {msg}{r}")

    print(f"\n{'=' * 60}")


def interactive_review(root: etree._Element) -> None:
    """CLI review loop — edit, delete, merge, then confirm."""
    print("\n" + "=" * 60)
    print(" REVIEW & CONFIRM")
    print("=" * 60)
    print("  e<id>  — Edit entry         (e.g. e3)")
    print("  x<id>  — Delete entry       (e.g. x3)")
    print("  v<id>  — View entry detail  (e.g. v3)")
    print("  m      — Merge duplicates")
    print("  s      — Proceed to SUBMIT  <-- when ready")
    print("=" * 60)

    while True:
        worklogs = root.findall("worklog")
        print_worklogs(worklogs)
        cmd = input("\n  Command (s=Submit / e<id> / x<id> / v<id> / m): ").strip().lower()

        if cmd in ("s", "q", "done", "submit", ""):
            break

        if cmd == "m":
            n = merge_duplicates(root)
            print(f"  [OK] Merged {n} duplicate entries.")
            continue

        action = cmd[0] if cmd else ""
        try:
            wid = int(cmd[1:])
        except (ValueError, IndexError):
            print("  [!] Unknown command. Try e3, x3, v3, m, or q.")
            continue

        wl = next((w for w in worklogs if w.get("id") == str(wid)), None)
        if wl is None:
            print(f"  [!] ID {wid} not found.")
            continue

        if action == "v":
            show_detail(wl)
        elif action == "e":
            edit_worklog(wl)
        elif action == "x":
            confirm = input(f"  Delete entry #{wid}? [y/N]: ").strip().lower()
            if confirm == "y":
                root.remove(wl)
                print(f"  [OK] Entry #{wid} deleted.")
        else:
            print("  [!] Unknown command. Try e3, x3, v3, m, or q.")


# ── Tempo submit ───────────────────────────────────────────────────────────────
def submit_to_tempo(
    root: etree._Element,
    config: dict,
    dry_run: bool = False,
) -> dict:
    worklogs = root.findall("worklog")
    submittable = [wl for wl in worklogs if wl.get("status") in ("OK", "Warning", "Edited — re-run validation")]
    results: dict = {"success": [], "failed": []}

    if not submittable:
        print("\n[!] No submittable worklogs found.")
        return results

    label = "[DRY RUN] " if dry_run else ""
    print(f"\n{label}Submitting {len(submittable)} worklog(s) to Tempo...\n")

    url = "https://api.tempo.io/4/worklogs"
    headers = {
        "Authorization": f"Bearer {config['tempo_token']}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }

    for wl in submittable:
        wid       = wl.get("id")
        issue_key = wl.findtext("issueKey") or ""
        issue_id  = wl.findtext("issueId") or ""
        date_str  = wl.findtext("date") or ""
        hours     = wl.findtext("hours") or "0"
        seconds   = int(wl.findtext("timeSpentSeconds") or 0)
        desc      = wl.findtext("description") or ""
        start_t   = wl.findtext("startTime") or "09:00:00"
        acct_id   = (wl.findtext("authorAccountId") or "").strip() or config.get("author_account_id", "")

        if not issue_id and not dry_run:
            results["failed"].append({
                "id": wid, "issue_key": issue_key, "date": date_str,
                "error": "Missing numeric issueId — Jira ID lookup failed for this key.",
                "suggestion": f"Check that {issue_key} exists in Jira and credentials are correct.",
            })
            print(f"  [SKIP] #{wid:<4} {issue_key:<22} — no numeric issueId, skipping.")
            continue

        payload = {
            "issueId":          int(issue_id) if issue_id else None,
            "timeSpentSeconds": seconds,
            "startDate":        date_str,
            "startTime":        start_t,
            "description":      desc,
            "authorAccountId":  acct_id,
        }

        if dry_run:
            stat  = wl.get("status", "OK")
            c     = _c(COLOR.get(stat, ""))
            r     = _c(COLOR["reset"])
            msgs  = get_validation_messages(wl)
            pfx   = STATUS_PREFIX.get(stat, stat[:5])
            print(f"  [DRY] #{wid:<4} {c}{pfx}{r}  {issue_key:<22} {date_str}  {hours}h  ->  {seconds}s")
            for msg in msgs:
                print(f"              {c}[!] {msg}{r}")
            results["success"].append({"id": wid, "issue_key": issue_key, "date": date_str,
                                       "warnings": msgs})
            continue

        try:
            resp = _session.post(url, json=payload, headers=headers, timeout=15)
            if resp.status_code in (200, 201):
                print(f"  [OK]   #{wid:<4} {issue_key:<22} {date_str}  {hours}h")
                results["success"].append({"id": wid, "issue_key": issue_key, "date": date_str})
            else:
                err_text = resp.text[:300]
                suggestion = "Check issue key, author accountId, and API token validity."
                if resp.status_code == 401:
                    suggestion = "Tempo token is invalid or expired."
                elif resp.status_code == 403:
                    suggestion = "No permission to log time on this issue."
                elif resp.status_code == 404:
                    suggestion = "Issue not found or Tempo plan does not include it."
                results["failed"].append({
                    "id": wid, "issue_key": issue_key, "date": date_str,
                    "error": f"HTTP {resp.status_code}: {err_text}",
                    "suggestion": suggestion,
                })
                print(f"  [FAIL] #{wid:<4} {issue_key:<22} → HTTP {resp.status_code}")
        except Exception as exc:
            results["failed"].append({
                "id": wid, "issue_key": issue_key, "date": date_str,
                "error": str(exc),
                "suggestion": "Check network connectivity and Tempo API URL.",
            })
            print(f"  [FAIL] #{wid:<4} {issue_key:<22} → {exc}")

    return results


# ── Report ─────────────────────────────────────────────────────────────────────
def print_report(results: dict, dry_run: bool = False) -> None:
    label = "DRY RUN -- " if dry_run else ""
    c_ok   = COLOR["OK"]
    c_warn = COLOR["Warning"]
    c_err  = COLOR["Error"]
    r      = COLOR["reset"]

    print(f"\n{'=' * 60}")
    print(f" {label}IMPORT REPORT")
    print(f"{'=' * 60}")

    verb = "would be created" if dry_run else "created"
    n_warn = sum(1 for s in results["success"] if s.get("warnings"))
    n_clean = len(results["success"]) - n_warn

    print(f"  {c_ok}OK     {r} {verb}: {n_clean} worklog(s)")
    print(f"  {c_warn}Warning{r} {verb}: {n_warn} worklog(s)  (submitted but had issues)")
    print(f"  {c_err}Error  {r} blocked:      {len(results['failed'])} worklog(s)  (NOT submitted)")

    if n_warn:
        print(f"\n  {c_warn}Warnings (submitted anyway):{r}")
        for s in results["success"]:
            if s.get("warnings"):
                print(f"    #{s['id']}  {s['issue_key']}  {s['date']}")
                for msg in s["warnings"]:
                    print(f"      ! {msg}")

    if results["failed"]:
        print(f"\n  {c_err}Failed entries:{r}")
        for f in results["failed"]:
            print(f"    #{f['id']}  {f['issue_key']}  {f['date']}")
            print(f"      Error:      {f['error'][:100]}")
            print(f"      Suggestion: {f.get('suggestion', '')}")


def export_errors_csv(results: dict, path: Path) -> None:
    if not results["failed"]:
        return
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=["id", "issue_key", "date", "error", "suggestion"])
        writer.writeheader()
        writer.writerows(results["failed"])
    print(f"\n  [INFO] Error report exported to: {path}")


# ── Verify ────────────────────────────────────────────────────────────────────
def _cmd_verify(config: dict, issue_keys: list[str], date_from: str, date_to: str) -> None:
    """Fetch worklogs from Tempo and display what is actually logged."""
    if config_is_placeholder(config):
        print("[!] Config has placeholder values — fill in tempo_token and jira credentials.")
        return

    tempo_hdrs = {
        "Authorization": f"Bearer {config['tempo_token']}",
        "Accept": "application/json",
    }

    # Build query params
    params: dict = {"limit": 1000}
    if date_from:
        params["from"] = date_from
    if date_to:
        params["to"] = date_to

    print(f"\n[*] Fetching worklogs from Tempo"
          f"{' for ' + ', '.join(issue_keys) if issue_keys else ''}"
          f"{' from ' + date_from if date_from else ''}"
          f"{' to ' + date_to if date_to else ''} ...")

    all_worklogs: list[dict] = []

    if issue_keys:
        # Fetch per issue key
        for ik in issue_keys:
            try:
                r = _session.get(
                    f"https://api.tempo.io/4/worklogs/issue/{ik}",
                    headers=tempo_hdrs, params=params, timeout=15,
                )
                if r.status_code == 200:
                    all_worklogs += r.json().get("results", [])
                elif r.status_code == 404:
                    print(f"  [!] Issue {ik}: no worklogs found (or issue does not exist in Tempo)")
                else:
                    print(f"  [!] Issue {ik}: HTTP {r.status_code} — {r.text[:100]}")
            except Exception as exc:
                print(f"  [!] Issue {ik}: {exc}")
    else:
        # Fetch all worklogs in date range for current user
        account_id = config.get("author_account_id", "")
        if account_id and not account_id.startswith("YOUR_"):
            params["accountId"] = account_id
        try:
            r = _session.get(
                "https://api.tempo.io/4/worklogs",
                headers=tempo_hdrs, params=params, timeout=15,
            )
            if r.status_code == 200:
                all_worklogs = r.json().get("results", [])
            else:
                print(f"  [!] Tempo API HTTP {r.status_code}: {r.text[:200]}")
                return
        except Exception as exc:
            print(f"  [!] Connection error: {exc}")
            return

    if not all_worklogs:
        print("\n  No worklogs found for the given criteria.")
        print("  Possible reasons:")
        print("    - No time was submitted yet (ran with --dry-run?)")
        print("    - Wrong date range (use --from / --to)")
        print("    - Wrong issue key (use --issue KEY)")
        print("    - author_account_id in config.json does not match the Tempo user")
        return

    # Print table
    total_seconds = 0
    print(f"\n  {'Issue Key':<22}  {'Date':<12}  {'Hours':>6}  {'Author':<25}  Description")
    print(f"  {'-'*22}  {'-'*12}  {'-'*6}  {'-'*25}  {'-'*40}")

    for wl in sorted(all_worklogs, key=lambda w: (w.get("startDate", ""), w.get("issue", {}).get("key", ""))):
        ik      = wl.get("issue", {}).get("key", "")
        dt      = wl.get("startDate", "")
        secs    = wl.get("timeSpentSeconds", 0)
        hrs     = secs / 3600
        author  = wl.get("author", {}).get("displayName", "")[:25]
        desc    = (wl.get("description") or "")[:40]
        total_seconds += secs
        print(f"  {ik:<22}  {dt:<12}  {hrs:>6.2f}  {author:<25}  {desc}")

    total_hrs = total_seconds / 3600
    print(f"\n  Total logged: {total_hrs:.2f}h  ({len(all_worklogs)} worklog entries)")


# ── Setup helpers ─────────────────────────────────────────────────────────────
def _cmd_whoami(config: dict) -> None:
    """Fetch and display the current Jira user's accountId."""
    if config_is_placeholder(config):
        print("[!] Config still has placeholder values.")
        print("    Fill in jira_url, jira_email, and jira_api_token in config.json first.")
        return

    base = config["jira_url"].rstrip("/")
    auth = (config["jira_email"], config["jira_api_token"])
    try:
        resp = _session.get(
            f"{base}/rest/api/3/myself",
            auth=auth,
            headers={"Accept": "application/json"},
            timeout=10,
        )
    except Exception as exc:
        print(f"[!] Connection failed: {exc}")
        return

    if resp.status_code != 200:
        print(f"[!] Jira returned HTTP {resp.status_code} — check jira_url and jira_api_token.")
        return

    data = resp.json()
    account_id  = data.get("accountId", "")
    display     = data.get("displayName", "")
    email       = data.get("emailAddress", "")

    print(f"\n  Jira user:  {display} <{email}>")
    print(f"  accountId:  {account_id}")
    print(f"\n  Copy the accountId value into config.json -> \"author_account_id\"")

    if account_id and not config_is_placeholder(config):
        update = input("\n  Update config.json with this accountId now? [y/N]: ").strip().lower()
        if update == "y":
            config["author_account_id"] = account_id
            Path("config.json").write_text(json.dumps(config, indent=2), encoding="utf-8")
            print("  [OK] config.json updated.")


# ── Main ────────────────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert Excel timesheets to XML and submit to Tempo Timesheets.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python timesheet_tool.py
  python timesheet_tool.py --dry-run
  python timesheet_tool.py --no-jira-check
  python timesheet_tool.py "JIRA EXPORT.xlsx" "RWE export.xlsx" --xml out.xml
"""
    )
    parser.add_argument("files", nargs="*", help="Input Excel files (default: auto-detect in current dir)")
    parser.add_argument("--xml",           default="worklogs.xml",  help="Output XML file (default: worklogs.xml)")
    parser.add_argument("--dry-run",       action="store_true",     help="Parse and validate only — do not write to Tempo")
    parser.add_argument("--no-jira-check", action="store_true",     help="Skip Jira issue existence validation")
    parser.add_argument("--whoami",        action="store_true",     help="Show your Jira account ID and exit")
    parser.add_argument("--verify",        action="store_true",     help="Show worklogs already logged in Tempo and exit")
    parser.add_argument("--issue",         default=None,            help="Issue key to verify (e.g. KAN-1). Repeatable: --issue KAN-1 --issue KAN-2", action="append")
    parser.add_argument("--from",          dest="date_from", default=None, help="Date filter start for --verify (YYYY-MM-DD)")
    parser.add_argument("--to",            dest="date_to",   default=None, help="Date filter end for --verify (YYYY-MM-DD)")
    parser.add_argument("--config",        default="config.json",   help="Path to config.json")
    parser.add_argument("--mapping",       default="mapping.json",  help="Path to mapping.json (WBS → Issue Key)")
    parser.add_argument("--remap",         default=None,            help="Path to key_mapping.json from setup_jira_issues.py")
    args = parser.parse_args()

    config  = load_config(Path(args.config))
    mapping = load_mapping(Path(args.mapping))

    # ── Configure SSL for corporate networks ────────────────────────────────────
    ssl_verify = config.get("ssl_verify", True)
    _session.verify = ssl_verify
    if not ssl_verify:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        print("[!] SSL verification disabled (ssl_verify=false in config.json).")

    if args.whoami:
        _cmd_whoami(config)
        sys.exit(0)

    if args.verify:
        _cmd_verify(config, args.issue or [], args.date_from or "", args.date_to or "")
        sys.exit(0)

    # ── Load key remap (old Excel keys → new Jira keys) ────────────────────────
    key_remap: dict[str, str] = {}
    if args.remap:
        remap_path = Path(args.remap)
        if not remap_path.exists():
            print(f"[!] Remap file not found: {remap_path}")
            sys.exit(1)
        remap_data = json.loads(remap_path.read_text(encoding="utf-8"))
        key_remap  = remap_data.get("key_mapping", {})
        # Also merge wbs/worklist mappings into main mapping
        mapping.setdefault("wbs_to_issue", {}).update(remap_data.get("wbs_mapping", {}))
        mapping.setdefault("worklist_to_issue", {}).update(remap_data.get("worklist_mapping", {}))
        print(f"[*] Key remap loaded: {len(key_remap)} issue key(s) remapped from {remap_path.name}")

    # ── Locate Excel files ──────────────────────────────────────────────────────
    if args.files:
        excel_paths = [Path(f) for f in args.files]
    else:
        excel_paths = [
            BASE_DIR / "JIRA EXPORT.xlsx",
            BASE_DIR / "RWE export.xlsx",
        ]

    all_entries: list[dict] = []
    for p in excel_paths:
        if not p.exists():
            print(f"[!] File not found: {p}")
            continue
        name_lower = p.name.lower()
        if "jira" in name_lower:
            print(f"[*] Parsing JIRA export:  {p.name}")
            all_entries += parse_jira_export(p)
        elif "rwe" in name_lower:
            print(f"[*] Parsing RWE export:   {p.name}")
            all_entries += parse_rwe_export(p, mapping)
        else:
            print(f"[!] Unknown format — skipping: {p.name}")

    if not all_entries:
        print("[!] No entries found. Exiting.")
        sys.exit(1)

    # Apply key remap (replace old Excel issue keys with new Jira keys)
    if key_remap:
        remapped = 0
        for entry in all_entries:
            old = entry.get("issue_key", "")
            if old in key_remap:
                entry["issue_key"] = key_remap[old]
                remapped += 1
        if remapped:
            print(f"[*] Remapped {remapped} worklog issue key(s) to new Jira keys.")

    print(f"[*] Parsed {len(all_entries)} entries total.")

    # ── Jira validation ─────────────────────────────────────────────────────────
    issue_validity: dict[str, bool] = {}
    if args.no_jira_check or config_is_placeholder(config):
        if config_is_placeholder(config):
            print("[!] Placeholder credentials detected — skipping Jira validation.")
            print("    Edit config.json and re-run to enable issue existence checks.")
        else:
            print("[*] Jira validation skipped (--no-jira-check).")
    else:
        unique_keys = {e["issue_key"] for e in all_entries if e.get("issue_key")}
        print(f"[*] Validating {len(unique_keys)} unique issue key(s) via Jira API...")
        issue_validity = validate_issues_via_jira(unique_keys, config)

    # ── Resolve numeric Jira issue IDs (required by Tempo API) ─────────────────
    issue_id_map: dict[str, int] = {}
    if not config_is_placeholder(config):
        unique_keys = {e["issue_key"] for e in all_entries if e.get("issue_key")}
        print(f"[*] Resolving numeric Jira IDs for {len(unique_keys)} issue key(s)...")
        issue_id_map = resolve_issue_ids(unique_keys, config)
        print(f"[*] Resolved {len(issue_id_map)}/{len(unique_keys)} issue IDs.")

    # ── Build XML ───────────────────────────────────────────────────────────────
    root = build_xml(all_entries, issue_validity, config, issue_id_map)

    # ── Show issues BEFORE anything else ────────────────────────────────────────
    print_issues_summary(root.findall("worklog"))

    xml_path = Path(args.xml)
    tree = etree.ElementTree(root)
    tree.write(str(xml_path), pretty_print=True, xml_declaration=True, encoding="utf-8")
    print(f"[*] XML written to: {xml_path}")

    # ── Review ──────────────────────────────────────────────────────────────────
    interactive_review(root)

    # Save edited XML
    tree.write(str(xml_path), pretty_print=True, xml_declaration=True, encoding="utf-8")

    # ── Confirm & submit ────────────────────────────────────────────────────────
    worklogs = root.findall("worklog")
    n_ok      = sum(1 for w in worklogs if w.get("status") == "OK")
    n_warn    = sum(1 for w in worklogs if w.get("status") == "Warning")
    n_edited  = sum(1 for w in worklogs if w.get("status") == "Edited — re-run validation")
    n_err     = sum(1 for w in worklogs if w.get("status") == "Error")
    n_submit  = n_ok + n_warn + n_edited

    if n_err:
        print(f"\n[!] {n_err} entries have Error status and will NOT be submitted.")
    if n_submit == 0:
        print("[!] Nothing to submit. Exiting.")
        sys.exit(0)

    if config_is_placeholder(config) and not args.dry_run:
        print("\n[!] Tempo token is a placeholder — forcing --dry-run mode.")
        args.dry_run = True

    mode_label = "DRY RUN" if args.dry_run else "SUBMIT TO TEMPO"
    confirm = input(
        f"\n[{mode_label}] Send {n_submit} worklog(s) to Tempo? [y/N]: "
    ).strip().lower()

    if confirm != "y":
        print("[*] Cancelled. XML saved for later use.")
        sys.exit(0)

    # ── Tempo call ──────────────────────────────────────────────────────────────
    results = submit_to_tempo(root, config, dry_run=args.dry_run)
    print_report(results, dry_run=args.dry_run)

    if results["failed"]:
        export_errors_csv(results, BASE_DIR / "import_errors.csv")

    tree.write(str(xml_path), pretty_print=True, xml_declaration=True, encoding="utf-8")
    print(f"\n[*] Final XML saved: {xml_path}")


if __name__ == "__main__":
    main()
